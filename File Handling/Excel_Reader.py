#
# File        : Excel_Reader.py  
# Created     : 25/10/2021
# Author      : James Bowman
# Version     : v1.0.0
# Licensing   : (C) James Bowman
#               Available under GNU Public License (GPL)
# Description : 
#
import openpyxl

def parse_xl_file(file_name):
    print("Hi")
    wb = openpyxl.load_workbook(file_name)
    ws = wb['Sheet1']
    print(ws['A1'].value)
    print(ws['A2'].value)

    for i in range(1, ws.max_row + 1):
        print("\n")
        print("Rows ", i, " data :")

        for j in range(1, ws.max_column + 1):
            cell_obj = ws.cell(row=i, column=j)
            print(cell_obj.value, end=" ")

if __name__ == '__main__':
    parse_xl_file('Financial_Sample.xlsx')
