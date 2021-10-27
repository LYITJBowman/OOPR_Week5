#
# File        : FH_Question1.py  
# Created     : 27/10/2021
# Author      : James Bowman
# Version     : v1.0.0
# Licensing   : (C) James Bowman
#               Available under GNU Public License (GPL)
# Description : 
#

import openpyxl

def tidy_xl_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb['Sheet1']

    # Strip Whitespace from first two columns
    for i in range(1, ws.max_row + 1):
        for j in range(1, 3):
            cell_obj = ws.cell(row=i, column=j)
            new_obj = str(cell_obj.value).strip()
            cell_obj.value = new_obj

    # UPPER One Column, just for the sake of it!
    for i in range(1, ws.max_row + 1):
        cell_obj = ws.cell(row=i, column=4)
        new_obj = str(cell_obj.value).upper()
        cell_obj.value = new_obj

    # Count the blanks per column
    for i in range(1, ws.max_row + 1):
        count = 0
        for j in range(1, ws.max_column + 1):
            cell_obj = ws.cell(row=i, column=j)
            if (cell_obj.value is None):
                count += 1
            print("Column {} has {} blank cells.".format(j, count))

    # Splitting Colunn 16
    for i in range(1, ws.max_row + 1):
        cell_obj_16 = ws.cell(row=i, column=16)
        cell_obj_17 = ws.cell(row=i, column=17)
        new_obj_16 = str(cell_obj_16.value[:2])
        new_obj_17 = str(cell_obj_16.value[2:])
        cell_obj_16.value = new_obj_16
        cell_obj_17.value = new_obj_17

    wb.save('Financials_Sample_New.xlsx')

if __name__ == '__main__':
    tidy_xl_file('Financial_Sample.xlsx')
